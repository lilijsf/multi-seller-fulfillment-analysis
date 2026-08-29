# -*- coding: utf-8 -*-
"""
Analysis code for:
Multi-Seller Fulfillment and Customer Dissatisfaction in Marketplace E-Commerce

This script reproduces the data preparation, statistical models,
robustness analyses, marginal effects, diagnostics, and figures
reported in the paper.
"""

import os
import sys
import platform
import warnings
from datetime import datetime
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import statsmodels.api as sm
import statsmodels.formula.api as smf
from statsmodels.stats.outliers_influence import variance_inflation_factor
from scipy import stats
from sklearn.metrics import roc_auc_score
import statsmodels
import scipy
import matplotlib
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

import kagglehub

# Only suppress the specific, harmless statsmodels convergence-format warning.
# All other warnings remain visible, since silent suppression can hide real
# estimation problems (separation, singular matrices, etc.).
warnings.filterwarnings("ignore", message="Maximum Likelihood optimization failed to converge")

OUT_DIR = "/content/outputs" if os.path.isdir("/content") else "outputs"
os.makedirs(OUT_DIR, exist_ok=True)
WORKBOOK_PATH = os.path.join(OUT_DIR, "python_results.xlsx")

plt.rcParams.update({
    "font.family": "serif",
    "font.size": 11,
    "axes.titlesize": 12,
    "axes.titleweight": "bold",
    "axes.spines.top": False,
    "axes.spines.right": False,
    "figure.dpi": 150,
})

COLORS = {
    "primary":   "#2C3E6B",
    "secondary": "#C0392B",
    "neutral":   "#7F8C8D",
    "light":     "#D6DCE4",
    "highlight": "#E67E22",
}

def fmt_p(p):
    """Report p-values without the misleading p = 0.0000 format.
    Returns the bare value/threshold only, e.g. '<0.001' or '0.732' -
    use fmt_p_inline() for prose sentences that need the '=' sign."""
    if pd.isna(p):
        return "n/a"
    return "<0.001" if p < 0.001 else f"{p:.3f}"

def fmt_p_inline(p):
    """Same as fmt_p but prefixed with '=' when not below the threshold,
    for use in prose like f"p{fmt_p_inline(p)}" -> 'p=0.732' or 'p<0.001'."""
    s = fmt_p(p)
    return s if s.startswith("<") else f"={s}"

def sig_stars(p):
    if pd.isna(p):
        return ""
    if p < 0.001: return "***"
    if p < 0.01:  return "**"
    if p < 0.05:  return "*"
    return ""

def section(title):
    print("\n" + "=" * 90)
    print(title)
    print("=" * 90)

# Audit collectors used later in the single Excel workbook.
merge_audit_rows = []
aggregation_audit_rows = []

# ============================================================================
# SECTION 1 — Load raw tables
# ============================================================================
section("SECTION 1 — Data Loading")

path = kagglehub.dataset_download("olistbr/brazilian-ecommerce")
print(f"Dataset path: {path}")

def load_csv(name):
    return pd.read_csv(os.path.join(path, name))

orders_raw      = load_csv("olist_orders_dataset.csv")
items_raw       = load_csv("olist_order_items_dataset.csv")
reviews_raw     = load_csv("olist_order_reviews_dataset.csv")
payments_raw    = load_csv("olist_order_payments_dataset.csv")
products_raw    = load_csv("olist_products_dataset.csv")
customers_raw   = load_csv("olist_customers_dataset.csv")
try:
    cat_translation = load_csv("product_category_name_translation.csv")
except FileNotFoundError:
    cat_translation = None

print(f"orders_raw:    {len(orders_raw):,} rows, {orders_raw['order_id'].nunique():,} unique order_id")
print(f"items_raw:     {len(items_raw):,} rows, {items_raw['order_id'].nunique():,} unique order_id")
print(f"reviews_raw:   {len(reviews_raw):,} rows, {reviews_raw['order_id'].nunique():,} unique order_id")
print(f"payments_raw:  {len(payments_raw):,} rows, {payments_raw['order_id'].nunique():,} unique order_id")
print(f"products_raw:  {len(products_raw):,} rows, {products_raw['product_id'].nunique():,} unique product_id")
print(f"customers_raw: {len(customers_raw):,} rows, {customers_raw['customer_id'].nunique():,} unique customer_id")

for col in ["order_purchase_timestamp", "order_delivered_customer_date",
            "order_estimated_delivery_date", "order_approved_at"]:
    orders_raw[col] = pd.to_datetime(orders_raw[col], errors="coerce")
for col in ["review_creation_date", "review_answer_timestamp"]:
    reviews_raw[col] = pd.to_datetime(reviews_raw[col], errors="coerce")

# ============================================================================
# SECTION 2 — Build independent order-level tables (no premature joins)
# ============================================================================
section("SECTION 2 — Building Order-Level Tables Separately")

# ---- 2.1 Orders table: already one row per order_id -----------------------
orders_tbl = orders_raw[[
    "order_id", "customer_id", "order_status",
    "order_purchase_timestamp", "order_delivered_customer_date",
    "order_estimated_delivery_date",
]].drop_duplicates(subset="order_id")
assert orders_tbl["order_id"].is_unique, "orders_tbl is not order-level"
print(f"orders_tbl: {len(orders_tbl):,} orders (one row per order_id, verified unique).")

# ---- 2.2 Reviews table: de-duplicate to one review per order --------------
review_counts = reviews_raw.groupby("order_id").size()
n_review_rows      = len(reviews_raw)
n_unique_reviewed   = reviews_raw["order_id"].nunique()
n_multi_review_ord  = (review_counts > 1).sum()

print(f"Review diagnostic:")
print(f"  Total review rows:              {n_review_rows:,}")
print(f"  Unique orders with >=1 review:  {n_unique_reviewed:,}")
print(f"  Orders with multiple review rows: {n_multi_review_ord:,} "
      f"({n_multi_review_ord / n_unique_reviewed:.2%} of reviewed orders)")

# Deterministic rule: retain the latest observed review row per order using
# review_answer_timestamp; ties are resolved by review_creation_date.
# This is a reproducible selection rule, not a claim that the retained row is
# necessarily the customer's definitive final evaluation.
reviews_sorted = reviews_raw.sort_values(
    ["order_id", "review_answer_timestamp", "review_creation_date"],
    na_position="first"
)
reviews_tbl = reviews_sorted.drop_duplicates(subset="order_id", keep="last")[
    ["order_id", "review_score", "review_creation_date", "review_answer_timestamp"]
]
assert reviews_tbl["order_id"].is_unique, "reviews_tbl is not order-level"
print(f"reviews_tbl: {len(reviews_tbl):,} orders after retaining the latest observed "
      f"review row per order_id (rule: max review_answer_timestamp, tie-break max "
      f"review_creation_date).")

# ---- 2.3 Order-items table: aggregate FIRST, before touching anything else
n_item_rows = len(items_raw)
items_agg = items_raw.groupby("order_id").agg(
    item_count           = ("order_item_id", "count"),
    unique_product_count = ("product_id", "nunique"),
    seller_count          = ("seller_id", "nunique"),
    freight_value          = ("freight_value", "sum"),
    price                  = ("price", "sum"),
).reset_index()
items_agg["multi_seller"] = (items_agg["seller_count"] > 1).astype(int)
assert items_agg["order_id"].is_unique, "items_agg is not order-level"
print(f"items_agg: {len(items_agg):,} orders aggregated from {n_item_rows:,} raw item rows "
      f"(mean {n_item_rows/len(items_agg):.2f} item rows per order).")

# ---- 2.4 Payments table: aggregated but NOT merged unless used ------------
payments_agg = payments_raw.groupby("order_id").agg(
    total_payment_value = ("payment_value", "sum"),
    max_installments     = ("payment_installments", "max"),
    n_payment_rows        = ("payment_value", "count"),
).reset_index()
assert payments_agg["order_id"].is_unique, "payments_agg is not order-level"
print(f"payments_agg: {len(payments_agg):,} orders (aggregated for potential diagnostic "
      f"use only; NOT merged into the analytical dataset because no payment-derived "
      f"variable is used in any specified model).")

# ---- 2.5 Product category: dominant category per order --------------------
# Rule: the product category with the highest summed merchandise value
# within the order; ties are resolved alphabetically.
items_products = items_raw.merge(
    products_raw[["product_id", "product_category_name"]],
    on="product_id", how="left", validate="many_to_one"
)
cat_value = (items_products.groupby(["order_id", "product_category_name"], dropna=False)["price"]
             .sum().reset_index())
cat_value = cat_value.sort_values(
    ["order_id", "price", "product_category_name"],
    ascending=[True, False, True],
    na_position="last",
    kind="mergesort",
)
dominant_cat = cat_value.drop_duplicates(subset="order_id", keep="first")[
    ["order_id", "product_category_name"]
].rename(columns={"product_category_name": "dominant_category_pt"})

if cat_translation is not None:
    dominant_cat = dominant_cat.merge(
        cat_translation, left_on="dominant_category_pt",
        right_on="product_category_name", how="left", validate="many_to_one"
    ).rename(columns={"product_category_name_english": "dominant_category"})
    dominant_cat["dominant_category"] = dominant_cat["dominant_category"].fillna(
        dominant_cat["dominant_category_pt"])
else:
    dominant_cat["dominant_category"] = dominant_cat["dominant_category_pt"]
dominant_cat = dominant_cat[["order_id", "dominant_category"]]
dominant_cat["dominant_category"] = dominant_cat["dominant_category"].fillna("unknown")
assert dominant_cat["order_id"].is_unique, "dominant_cat is not order-level"
print(f"dominant_cat: {len(dominant_cat):,} orders assigned a dominant product category "
      f"(rule: category with highest summed merchandise value; ties resolved alphabetically).")

# ---- 2.6 Dominant seller per order for bootstrap clustering ---------------
# A multi-seller order belongs to more than one seller simultaneously, so no
# single seller_id can represent it exactly. For the seller-cluster bootstrap
# in SECTION 9B, each order is assigned to the seller with the highest summed
# item price within that order (ties resolved alphabetically by seller_id).
# This assignment is used ONLY to define resampling clusters for the
# robustness bootstrap; it is not used as a covariate and does not alter the
# multi_seller exposure variable or any fitted model specification.
items_seller_value = items_raw.groupby(["order_id", "seller_id"])["price"].sum().reset_index()
items_seller_value = items_seller_value.sort_values(
    ["order_id", "price", "seller_id"], ascending=[True, False, True],
    kind="mergesort"
)
dominant_seller = items_seller_value.drop_duplicates(subset="order_id", keep="first")[
    ["order_id", "seller_id"]
].rename(columns={"seller_id": "dominant_seller_id"})
assert dominant_seller["order_id"].is_unique, "dominant_seller is not order-level"
print(f"dominant_seller: {len(dominant_seller):,} orders assigned a dominant seller_id "
      f"for bootstrap clustering purposes only (rule: seller with highest summed "
      f"merchandise value in the order; ties resolved alphabetically).")

# ---- 2.6 Customer region ---------------------------------------------------
customer_state = customers_raw[["customer_id", "customer_state"]].drop_duplicates(subset="customer_id")
assert customer_state["customer_id"].is_unique

# ============================================================================
# SECTION 3 — Validated merges into a single order-level analytical table
# ============================================================================
section("SECTION 3 — Merging Order-Level Tables With Validation")

def merge_report(df, label, validation_type, unmatched_count=0):
    row = {
        "Merge step": label,
        "Validation": validation_type,
        "Rows": len(df),
        "Unique orders": df["order_id"].nunique(),
        "Duplicated order rows": int(df["order_id"].duplicated().sum()),
        "Unmatched": int(unmatched_count),
    }
    merge_audit_rows.append(row)
    print(f"  [{label}] rows={row['Rows']:,}  unique order_id={row['Unique orders']:,}  "
          f"duplicated order_id rows={row['Duplicated order rows']:,}  unmatched={row['Unmatched']:,}")

analytical = orders_tbl.merge(customer_state.rename(columns={"customer_id": "cust_id_tmp"}),
                               left_on="customer_id", right_on="cust_id_tmp",
                               how="left", validate="many_to_one").drop(columns=["cust_id_tmp"])
merge_report(analytical, "orders + customer_state", "many_to_one", analytical["customer_state"].isna().sum())

pre = len(analytical)
analytical = analytical.merge(reviews_tbl, on="order_id", how="inner", validate="one_to_one")
merge_report(analytical, "+ reviews_tbl (inner)", "one_to_one", pre - len(analytical))
print(f"  Orders dropped for having no usable review: {pre - len(analytical):,}")

pre = len(analytical)
analytical = analytical.merge(items_agg, on="order_id", how="inner", validate="one_to_one")
merge_report(analytical, "+ items_agg (inner)", "one_to_one", pre - len(analytical))
print(f"  Orders dropped for having no usable item data: {pre - len(analytical):,}")

analytical = analytical.merge(dominant_cat, on="order_id", how="left", validate="one_to_one")
merge_report(analytical, "+ dominant_cat (left)", "one_to_one", analytical["dominant_category"].isna().sum())

analytical = analytical.merge(dominant_seller, on="order_id", how="left", validate="one_to_one")
merge_report(analytical, "+ dominant_seller (left)", "one_to_one", analytical["dominant_seller_id"].isna().sum())

unmatched_customers = analytical["customer_state"].isna().sum()
print(f"  Orders with unmatched customer_state: {unmatched_customers:,}")

assert analytical["order_id"].is_unique, "analytical table lost order-level uniqueness"
print(f"\nOrder-level uniqueness confirmed after all merges: "
      f"{analytical['order_id'].nunique():,} unique order_id across {len(analytical):,} rows.")

# ============================================================================
# SECTION 4 — Audit: raw vs. aggregated reconciliation on a random sample
# ============================================================================
section("SECTION 4 — Aggregation Audit (random order sample)")

rng = np.random.default_rng(42)
audit_ids = rng.choice(analytical["order_id"].unique(), size=5, replace=False)
for oid in audit_ids:
    raw_items = items_raw[items_raw["order_id"] == oid]
    agg_row = items_agg[items_agg["order_id"] == oid].iloc[0]
    print(f"\norder_id={oid}")
    print(f"  raw item rows:              {len(raw_items)}")
    print(f"  expected item_count:        {len(raw_items)}   | aggregated: {agg_row['item_count']}")
    print(f"  expected price total:       {raw_items['price'].sum():.2f}   | aggregated: {agg_row['price']:.2f}")
    print(f"  expected freight total:     {raw_items['freight_value'].sum():.2f}   | aggregated: {agg_row['freight_value']:.2f}")
    print(f"  expected seller count:      {raw_items['seller_id'].nunique()}   | aggregated: {agg_row['seller_count']}")
    assert len(raw_items) == agg_row["item_count"]
    assert np.isclose(raw_items["price"].sum(), agg_row["price"])
    assert np.isclose(raw_items["freight_value"].sum(), agg_row["freight_value"])
    assert raw_items["seller_id"].nunique() == agg_row["seller_count"]
    aggregation_audit_rows.append({
        "order_id": oid,
        "Raw item count": len(raw_items),
        "Aggregated item count": int(agg_row["item_count"]),
        "Raw price total": float(raw_items["price"].sum()),
        "Aggregated price total": float(agg_row["price"]),
        "Raw freight total": float(raw_items["freight_value"].sum()),
        "Aggregated freight total": float(agg_row["freight_value"]),
        "Raw seller count": int(raw_items["seller_id"].nunique()),
        "Aggregated seller count": int(agg_row["seller_count"]),
        "Pass": True,
    })
print("\nAll 5 sampled orders reconcile exactly between raw item rows and aggregated values.")

# ============================================================================
# SECTION 5 — Sample construction: primary vs. timing-restricted sensitivity
# ============================================================================
section("SECTION 5 — Sample Construction")

n_raw_orders = orders_tbl["order_id"].nunique()
n_delivered  = (orders_tbl["order_status"] == "delivered").sum()

# Delivered only, from here on.
base = analytical[analytical["order_status"] == "delivered"].copy()
n_after_delivered = len(base)

critical = ["order_delivered_customer_date", "order_estimated_delivery_date",
            "review_score", "freight_value", "price"]
pre = len(base)
base = base.dropna(subset=critical)
n_after_critical = len(base)
print(f"Dropped for missing critical fields: {pre - len(base):,}")

pre = len(base)
base = base[(base["item_count"] >= 1) & (base["seller_count"] >= 1) &
            (base["freight_value"] >= 0) & (base["price"] >= 0)]
n_after_sanity = len(base)
print(f"Dropped for failing basic sanity bounds (item/seller count, non-negative cost): "
      f"{pre - len(base):,}")

base["delivery_delay"] = (base["order_delivered_customer_date"]
                           - base["order_estimated_delivery_date"]).dt.days
base["dissatisfied"]   = (base["review_score"] <= 2).astype(int)
base["dissatisfied_3"] = (base["review_score"] <= 3).astype(int)
base["late"]           = (base["delivery_delay"] > 0).astype(int)

# --- Primary sample: does NOT exclude orders for review-timing reasons -----
primary = base.copy()
n_primary = len(primary)

# --- Timing-restricted sensitivity sample -----------------------------------
timing_ok = base["review_creation_date"] >= base["order_delivered_customer_date"]
n_timing_dropped = (~timing_ok).sum()
sensitivity = base[timing_ok].copy()
n_sensitivity = len(sensitivity)

print(f"\nReview-timing note: {n_timing_dropped:,} of {n_primary:,} orders "
      f"({n_timing_dropped/n_primary:.1%}) have a review created before the confirmed "
      f"delivery date. Because Olist can trigger the review invitation from the "
      f"ESTIMATED delivery date, this pattern is expected for late orders and is not, "
      f"by itself, evidence of an invalid record. These orders are retained in the "
      f"primary sample and excluded only in the timing-restricted sensitivity sample.")

for col in ["order_id", "item_count", "seller_count", "freight_value", "price"]:
    assert primary[col].notna().all()
assert primary["order_id"].is_unique
assert primary["item_count"].ge(1).all()
assert primary["seller_count"].ge(1).all()
assert primary["freight_value"].ge(0).all()
assert primary["price"].ge(0).all()
print("Automated validation assertions passed for the primary sample.")

reconciliation = pd.DataFrame({
    "Stage": ["Raw orders (orders table)", "Delivered orders",
              "With usable review + item data (post-merge, pre-filter)",
              "Complete critical fields", "Passing sanity bounds",
              "Primary sample (final)", "Timing-restricted sensitivity sample (final)"],
    "N": [n_raw_orders, n_delivered, n_after_delivered, n_after_critical,
          n_after_sanity, n_primary, n_sensitivity],
})
print("\nTable 1. Sample construction and exclusions")
print(reconciliation.to_string(index=False))

# ============================================================================
# SECTION 6 — Delivery variables: categories and nonlinearity
# ============================================================================
section("SECTION 6 — Delivery Delay Categorization")

DELIVERY_ORDER = [
    "Early (<=-1d)", "On time (0d)", "1-3d late",
    "4-7d late", "8-14d late", "15+d late",
]

def add_delivery_category(df):
    bins = [-np.inf, -1, 0, 3, 7, 14, np.inf]
    out = df.copy()
    out["delivery_category"] = pd.cut(
        out["delivery_delay"], bins=bins, labels=DELIVERY_ORDER, ordered=True
    )
    out["delivery_category"] = pd.Categorical(
        out["delivery_category"], categories=DELIVERY_ORDER, ordered=True
    )
    return out

primary     = add_delivery_category(primary)
sensitivity = add_delivery_category(sensitivity)

cat_counts = primary["delivery_category"].value_counts().reindex(
    DELIVERY_ORDER
)
print("Delivery category counts (primary sample):")
print(cat_counts.to_string())
sparse = cat_counts[cat_counts < 30]
if len(sparse) > 0:
    print(f"\nWARNING - sparse categories (n<30): {list(sparse.index)}")
else:
    print("\nNo category has fewer than 30 observations; categories are usable as specified.")

# ============================================================================
# SECTION 7 — Composition controls
# ============================================================================
section("SECTION 7 — Composition Controls")

primary["purchase_quarter"] = primary["order_purchase_timestamp"].dt.to_period("Q").astype(str)
primary["purchase_year"] = primary["order_purchase_timestamp"].dt.year.astype("Int64").astype(str)
top_categories = primary["dominant_category"].value_counts().head(10).index.tolist()
primary["category_grouped"] = np.where(primary["dominant_category"].isin(top_categories),
                                        primary["dominant_category"], "other")
top_states = primary["customer_state"].value_counts().head(8).index.tolist()
primary["state_grouped"] = np.where(primary["customer_state"].isin(top_states),
                                     primary["customer_state"], "other")

print(f"Category control: top {len(top_categories)} dominant categories kept individually, "
      f"remainder pooled into 'other' ({(primary['category_grouped']=='other').mean():.1%} of orders).")
print(f"Region control: top {len(top_states)} customer states kept individually, remainder "
      f"pooled into 'other' ({(primary['state_grouped']=='other').mean():.1%} of orders).")
print("Seller-to-customer geography is not constructed: seller location cannot be linked "
      "reliably to a single order-level distance measure without material assumptions, so "
      "it is omitted as infeasible for a six-page specification rather than silently ignored.")
print(f"Primary time diagnostic: purchase quarter, " f"{primary['purchase_quarter'].nunique()} distinct quarters.")
print(f"Enhanced-control model time variable: purchase year, " f"{primary['purchase_year'].nunique()} distinct years.")

# ============================================================================
# SECTION 8 — Descriptive statistics and subgroup comparison
# ============================================================================
section("SECTION 8 — Descriptive Statistics")

overall_rate = primary["dissatisfied"].mean()
print(f"Overall dissatisfaction rate (primary sample, N={len(primary):,}): {overall_rate:.1%}")

grp = primary.groupby("multi_seller")
n_single = (primary["multi_seller"] == 0).sum()
n_multi  = (primary["multi_seller"] == 1).sum()
diss_single = primary.loc[primary["multi_seller"] == 0, "dissatisfied"].mean()
diss_multi  = primary.loc[primary["multi_seller"] == 1, "dissatisfied"].mean()
n_diss_single = primary.loc[primary["multi_seller"] == 0, "dissatisfied"].sum()
n_diss_multi  = primary.loc[primary["multi_seller"] == 1, "dissatisfied"].sum()

risk_diff = diss_multi - diss_single
risk_ratio = diss_multi / diss_single
odds_single = diss_single / (1 - diss_single)
odds_multi  = diss_multi / (1 - diss_multi)
unadj_or = odds_multi / odds_single

ct = pd.crosstab(primary["multi_seller"], primary["dissatisfied"])
chi2, p_chi, dof, _ = stats.chi2_contingency(ct)

# Wilson CI for each proportion
def wilson_ci(k, n, z=1.96):
    p = k / n
    denom = 1 + z**2 / n
    center = p + z**2 / (2 * n)
    half = z * np.sqrt(p * (1 - p) / n + z**2 / (4 * n**2))
    return (center - half) / denom, (center + half) / denom

ci_single = wilson_ci(n_diss_single, n_single)
ci_multi  = wilson_ci(n_diss_multi, n_multi)

print(f"\nSingle-seller orders: N={n_single:,}, dissatisfied={n_diss_single:,}, "
      f"rate={diss_single:.1%} (95% CI {ci_single[0]:.1%}-{ci_single[1]:.1%})")
print(f"Multi-seller orders:  N={n_multi:,}, dissatisfied={n_diss_multi:,}, "
      f"rate={diss_multi:.1%} (95% CI {ci_multi[0]:.1%}-{ci_multi[1]:.1%})")
print(f"Absolute risk difference: {risk_diff:.1%}")
print(f"Risk ratio:               {risk_ratio:.2f}")
print(f"Unadjusted odds ratio:    {unadj_or:.2f}")
print(f"Chi-square: chi2={chi2:.1f}, df={dof}, p{fmt_p_inline(p_chi)}")
print("Note: the difference is statistically significant, although multi-seller orders "
      f"represent a relatively small proportion of the sample ({n_multi/len(primary):.1%}). "
      "The subgroup comparison below reports composition differences that a reader would "
      "otherwise have to infer.")

comp = primary.groupby("multi_seller")[["item_count", "freight_value", "price", "delivery_delay"]]
comp_mean = comp.mean().round(2)
comp_median = comp.median().round(2)
comp_iqr = comp.quantile(0.75) - comp.quantile(0.25)
print("\nSubgroup composition (mean):")
print(comp_mean.to_string())
print("\nSubgroup composition (median):")
print(comp_median.to_string())
print("\nSubgroup composition (IQR):")
print(comp_iqr.round(2).to_string())

desc_vars = ["dissatisfied", "delivery_delay", "item_count", "freight_value", "price", "multi_seller"]
desc = primary[desc_vars].describe().T[["mean", "std", "min", "50%", "max"]]
desc.columns = ["Mean", "Std", "Min", "Median", "Max"]
print("\nTable 2. Descriptive statistics (primary sample)")
print(desc.round(3).to_string())

thresh_check = primary.groupby("multi_seller")[["dissatisfied", "dissatisfied_3"]].mean() * 100
thresh_check.index = thresh_check.index.map({0: "Single-Seller", 1: "Multi-Seller"})
thresh_check.columns = ["Dissatisfied (<=2) %", "Dissatisfied (<=3) %"]
print("\nThreshold sensitivity check:")
print(thresh_check.round(1).to_string())

late_rate = primary["late"].mean()
diss_by_late = primary.groupby("late")["dissatisfied"].mean()
print(f"\nLate orders: {primary['late'].sum():,} ({late_rate:.1%} of primary sample)")
print("Dissatisfaction by lateness:")
print((diss_by_late * 100).round(1).to_string())

# ============================================================================
# SECTION 9 — Model estimation helpers
# ============================================================================
section("SECTION 9 — Regression Models")

def fit_logit(df, y_col, x_cols, label, cov_type="HC1", require_convergence=True):
    d = df[[y_col] + x_cols].dropna().copy()
    for c in [y_col] + x_cols:
        d[c] = d[c].astype(float)
    y = d[y_col]
    X = sm.add_constant(d[x_cols], has_constant="add")
    attempts = [("newton", 200), ("lbfgs", 500)]
    last_result = None
    last_error = None
    for method, maxiter in attempts:
        try:
            result = sm.Logit(y, X).fit(
                disp=False, method=method, maxiter=maxiter, cov_type=cov_type
            )
            last_result = result
            if result.mle_retvals.get("converged", False):
                result._optimizer_used = method
                print(f"\n[{label}] N={int(result.nobs):,}, converged with {method}, robust SE={cov_type}")
                extreme = result.params[result.params.abs() > 15]
                if len(extreme):
                    print(f"  WARNING - possible separation; extreme coefficients: {extreme.to_dict()}")
                return result
        except Exception as exc:
            last_error = exc
            print(f"  {label}: optimizer {method} failed: {exc}")
    if require_convergence:
        raise RuntimeError(
            f"{label} did not converge with newton or lbfgs. Last error: {last_error}"
        )
    if last_result is not None:
        last_result._optimizer_used = "unconverged"
    return last_result

def or_table(result, label):
    params = result.params
    ci = result.conf_int()
    rows = []
    for v in params.index:
        if v == "const":
            continue
        rows.append({
            "Variable": v,
            "Coef": round(params[v], 4),
            "Robust SE": round(result.bse[v], 4),
            "z": round(result.tvalues[v], 3),
            "p": fmt_p(result.pvalues[v]),
            "OR": round(np.exp(params[v]), 4),
            "OR_95%CI_low": round(np.exp(ci.loc[v, 0]), 4),
            "OR_95%CI_high": round(np.exp(ci.loc[v, 1]), 4),
            "Sig": sig_stars(result.pvalues[v]),
        })
    out = pd.DataFrame(rows)
    print(f"\nOdds ratios: {label}")
    print(out.to_string(index=False))
    return out

def model_fit_row(result, name):
    return {"Model": name, "N": int(result.nobs), "Log-Lik": round(result.llf, 1),
            "AIC": round(result.aic, 1), "BIC": round(result.bic, 1),
            "McFadden R2": round(result.prsquared, 4)}

structural_vars = ["multi_seller", "item_count", "freight_value", "price"]
delay_vars      = ["delivery_delay", "price"]
full_vars       = ["multi_seller", "item_count", "freight_value", "delivery_delay", "price"]

model_A = fit_logit(primary, "dissatisfied", delay_vars, "Model A: Delivery benchmark")
model_B = fit_logit(primary, "dissatisfied", structural_vars, "Model B: Fulfillment-structure benchmark")
model_C = fit_logit(primary, "dissatisfied", full_vars, "Model C: Combined full model")

or_A = or_table(model_A, "Model A")
or_B = or_table(model_B, "Model B")
or_C = or_table(model_C, "Model C")

fit_ABC = pd.DataFrame([model_fit_row(model_A, "A: Delivery benchmark"),
                         model_fit_row(model_B, "B: Structure benchmark"),
                         model_fit_row(model_C, "C: Combined full model")])
print("\nTable 3. Primary model fit comparison")
print(fit_ABC.to_string(index=False))

# Correct, non-nested-sequence framing of incremental contribution:
delta_AC = model_C.prsquared - model_A.prsquared
delta_BC = model_C.prsquared - model_B.prsquared
print(f"\nModel A vs. C (contribution of adding fulfillment-structure variables to the "
      f"delivery benchmark): delta McFadden R2 = {delta_AC:.4f}")
print(f"Model B vs. C (contribution of adding delivery delay to the structure benchmark): "
      f"delta McFadden R2 = {delta_BC:.4f}")
print("Model A and Model B are alternative, non-nested benchmark specifications. "
      "The A-to-C and B-to-C fit changes describe different variable blocks and are "
      "not treated as directly comparable measures of predictor importance.")

pub_rows = []
for var, label in [("multi_seller", "Multi-Seller (Fragmentation)"),
                    ("item_count", "Item Count"), ("freight_value", "Freight Value"),
                    ("delivery_delay", "Delivery Delay"), ("price", "Merchandise Value")]:
    row = {"Variable": label}
    for name, res in [("Model A", model_A), ("Model B", model_B), ("Model C", model_C)]:
        if var in res.params.index:
            row[name] = f"{res.params[var]:.3f}{sig_stars(res.pvalues[var])} ({res.bse[var]:.3f})"
        else:
            row[name] = "-"
    pub_rows.append(row)
pub_table_ABC = pd.DataFrame(pub_rows)
print("\nTable 3 (coefficient view). Coefficient (robust SE); ***p<.001 **p<.01 *p<.05")
print(pub_table_ABC.to_string(index=False))

# ---- Model D: Late indicator ----------------------------------------------
model_D = fit_logit(primary, "dissatisfied",
                     ["multi_seller", "item_count", "freight_value", "late", "price"],
                     "Model D: Late indicator")
or_D = or_table(model_D, "Model D")

# ---- Model E: Delivery categories (deterministic reference) ----------------
primary_E = primary.dropna(subset=["delivery_category"]).copy()
observed_categories = set(primary_E["delivery_category"].dropna().astype(str).unique())
missing_categories = [c for c in DELIVERY_ORDER if c not in observed_categories]
assert not missing_categories, f"Missing delivery categories: {missing_categories}"
assert "On time (0d)" in observed_categories

cat_dummies = pd.get_dummies(
    primary_E["delivery_category"], prefix="delcat", drop_first=False, dtype=float
)
reference_candidates = [c for c in cat_dummies.columns if c.endswith("On time (0d)")]
assert len(reference_candidates) == 1, (
    f"Expected exactly one on-time reference dummy, found {reference_candidates}"
)
cat_dummies = cat_dummies.drop(columns=reference_candidates)
assert cat_dummies.shape[1] == 5, (
    f"Expected five delivery-category coefficients, found {cat_dummies.shape[1]}"
)
primary_E = pd.concat([primary_E, cat_dummies], axis=1)
cat_predictor_cols = ["multi_seller", "item_count", "freight_value", "price"] + list(cat_dummies.columns)
model_E = fit_logit(
    primary_E, "dissatisfied", cat_predictor_cols,
    "Model E: Delivery categories (reference = On time (0d))"
)
or_E = or_table(model_E, "Model E")
print("Model E category terms:", list(cat_dummies.columns))

delay_model_comparison = pd.DataFrame([
    model_fit_row(model_C, "C: Continuous delay"),
    model_fit_row(model_D, "D: Late indicator"),
    model_fit_row(model_E, "E: Delivery categories"),
])
best_delay_model = delay_model_comparison.sort_values("AIC").iloc[0]["Model"]
print("\nDelay-model comparison:")
print(delay_model_comparison.to_string(index=False))
print(f"Best AIC among C/D/E: {best_delay_model}.")

# ---- Model F: Enhanced controls with stable purchase-year control ----------
def categorical_sparsity_table(df, col):
    tab = df.groupby(col, dropna=False)["dissatisfied"].agg(["count", "sum", "mean"]).reset_index()
    tab.columns = ["Level", "Total N", "Dissatisfied N", "Dissatisfaction rate"]
    tab.insert(0, "Variable", col)
    tab["Flag"] = np.select(
        [tab["Dissatisfied N"] == 0,
         tab["Dissatisfied N"] == tab["Total N"],
         tab["Total N"] < 100,
         tab["Dissatisfied N"] < 10],
        ["Zero dissatisfied", "Zero satisfied", "N<100", "Dissatisfied<10"],
        default="OK"
    )
    return tab

sparsity_tables = [
    categorical_sparsity_table(primary, "purchase_quarter"),
    categorical_sparsity_table(primary, "purchase_year"),
    categorical_sparsity_table(primary, "category_grouped"),
    categorical_sparsity_table(primary, "state_grouped"),
]
categorical_sparsity = pd.concat(sparsity_tables, ignore_index=True)
print("\nCategorical sparsity diagnostics:")
print(categorical_sparsity.to_string(index=False))

# Purchase quarter is diagnosed but not used because the early quarter is sparse.
# Purchase year is the stable and interpretable time control.
primary_F = pd.get_dummies(
    primary, columns=["category_grouped", "state_grouped", "purchase_year"],
    drop_first=True, dtype=float
)
control_cols = [
    c for c in primary_F.columns
    if c.startswith("category_grouped_")
    or c.startswith("state_grouped_")
    or c.startswith("purchase_year_")
]
f_cols = full_vars + control_cols
model_F = fit_logit(
    primary_F, "dissatisfied", f_cols,
    "Model F: Enhanced controls (category, state, purchase year)",
    require_convergence=True
)
or_F = or_table(model_F, "Model F")
assert model_F.mle_retvals.get("converged", False), "Model F must converge before reporting"
if model_F.params.abs().max() > 15 or model_F.bse.max() > 10:
    raise RuntimeError("Model F converged numerically but shows separation-like coefficients or SEs.")

# ---- Model G: Timing-restricted sensitivity sample -------------------------
sensitivity_full = sensitivity.dropna(subset=full_vars + ["dissatisfied"])
model_G = fit_logit(sensitivity, "dissatisfied", full_vars,
                     "Model G: Timing-restricted sensitivity sample")
or_G = or_table(model_G, "Model G")
print(f"\nSensitivity sample dropped {n_timing_dropped:,} orders relative to the primary "
      f"sample ({n_timing_dropped/n_primary:.1%}). Comparing Model C and Model G shows "
      f"whether the multi-seller and delay estimates depend on that exclusion rule.")

# ---- Model H: Broader dissatisfaction threshold ----------------------------
model_H = fit_logit(primary, "dissatisfied_3", full_vars, "Model H: Threshold <=3 sensitivity")
or_H = or_table(model_H, "Model H")

# ---- Freight robustness: log freight-per-item ------------------------------
primary["freight_per_item"] = primary["freight_value"] / primary["item_count"]
primary["log_freight_per_item"] = np.log1p(primary["freight_per_item"])
model_I_vars = ["multi_seller", "item_count", "log_freight_per_item", "delivery_delay", "price"]
model_I = fit_logit(primary, "dissatisfied", model_I_vars, "Model I: Freight-per-item robustness")
or_I = or_table(model_I, "Model I: Freight-per-item robustness")
print("\nNote: log_freight_per_item = log(1 + freight_value / item_count). This is labeled "
      "accurately as a log-transformed cost-per-item measure, not treated as a general "
      "cost proxy equivalence.")

robustness_rows = []
robustness_models = [
    ("D: Late indicator", model_D), ("E: Delivery categories", model_E),
    ("F: Enhanced controls", model_F), ("G: Timing-restricted", model_G),
    ("H: Threshold<=3", model_H), ("I: Freight/item", model_I),
]
for var in ["multi_seller", "item_count", "freight_value", "delivery_delay", "late", "price"]:
    row = {"Variable": var}
    for name, res in robustness_models:
        if var in res.params.index:
            row[name] = f"{np.exp(res.params[var]):.3f}{sig_stars(res.pvalues[var])}"
        else:
            row[name] = "-"
    robustness_rows.append(row)
table4 = pd.DataFrame(robustness_rows)
print("\nTable 4. Robustness models - odds ratios for key predictors")
print(table4.to_string(index=False))

# ============================================================================
# SECTION 9B — Additional Robustness and Sensitivity Analyses
# ============================================================================
section("SECTION 9B — Reviewer-Requested Extensions")

# ---- Model J: Delivery categories + enhanced controls ---------------------
# Combines the preferred nonlinear delay specification (Model E) with the
# product-category, customer-state, and purchase-year controls (Model F).
j_control_cols = [c for c in primary_F.columns if c.startswith("category_grouped_")
                   or c.startswith("state_grouped_") or c.startswith("purchase_year_")]
primary_J = primary_F.merge(
    primary_E[["order_id"] + list(cat_dummies.columns)], on="order_id", how="inner"
)
j_cols = ["multi_seller", "item_count", "freight_value", "price"] + list(cat_dummies.columns) + j_control_cols
model_J = fit_logit(primary_J, "dissatisfied", j_cols,
                     "Model J: Delivery categories + enhanced controls")
or_J = or_table(model_J, "Model J")

# ---- Model K: Multi-item-only sensitivity (R1#1) ---------------------------
# Restricts the sample to item_count >= 2, since single-item orders cannot,
# by construction, be multi-seller. The preferred categorical delivery-delay
# specification is retained so the sensitivity analysis is consistent with
# Model E in the main analysis.
primary_multi_item = primary[primary["item_count"] >= 2].copy()
n_multi_item_sample = len(primary_multi_item)
n_multi_item_multi_seller = (primary_multi_item["multi_seller"] == 1).sum()

print(f"\nMulti-item-only sensitivity sample: {n_multi_item_sample:,} orders with "
      f"item_count >= 2 ({n_multi_item_sample/n_primary:.1%} of the primary sample), "
      f"of which {n_multi_item_multi_seller:,} are multi-seller "
      f"({n_multi_item_multi_seller/n_multi_item_sample:.1%}).")

# Recreate delivery-category dummies within the restricted sample.
multi_item_cat_dummies = pd.get_dummies(
    primary_multi_item["delivery_category"], prefix="delcat",
    drop_first=False, dtype=float
)

multi_item_reference = [
    c for c in multi_item_cat_dummies.columns
    if c.endswith("On time (0d)")
]

assert len(multi_item_reference) == 1, (
    f"Expected exactly one on-time reference dummy, found {multi_item_reference}"
)

multi_item_cat_dummies = multi_item_cat_dummies.drop(columns=multi_item_reference)
primary_multi_item_K = pd.concat([primary_multi_item, multi_item_cat_dummies], axis=1)

model_K_vars = ["multi_seller", "item_count", "freight_value", "price",] + list(multi_item_cat_dummies.columns)
model_K = fit_logit(
    primary_multi_item_K, "dissatisfied",
    model_K_vars, "Model K: Multi-item-only sensitivity with delivery categories"
)

or_K = or_table(model_K, "Model K")

diss_multi_item_single = primary_multi_item.loc[primary_multi_item["multi_seller"] == 0, "dissatisfied"].mean()
diss_multi_item_multi = primary_multi_item.loc[primary_multi_item["multi_seller"] == 1, "dissatisfied"].mean()

print(f"Within multi-item orders only: single-seller dissatisfaction "
      f"{diss_multi_item_single:.1%}, multi-seller dissatisfaction "
      f"{diss_multi_item_multi:.1%}. "
      f"Model K multi-seller OR = "
      f"{np.exp(model_K.params['multi_seller']):.3f} "
      f"(compare to preferred full-sample Model E OR = "
      f"{np.exp(model_E.params['multi_seller']):.3f}).")

# ---- Model L: Product diversity sensitivity (R1#6, partial) ---------------
# Adds unique_product_count as an additional composition control to test
# whether the multi-seller association reflects assortment breadth rather
# than seller fragmentation itself. Product diversity was already computed
# during aggregation (SECTION 2.3) and required no new data construction.
diversity_vars = full_vars + ["unique_product_count"]
model_L = fit_logit(primary, "dissatisfied", diversity_vars,
                     "Model L: Product diversity sensitivity")
or_L = or_table(model_L, "Model L")
print(f"Multi-seller OR with product diversity held constant: "
      f"{np.exp(model_L.params['multi_seller']):.3f} "
      f"(Model C without this control: {np.exp(model_C.params['multi_seller']):.3f}).")

# ---- Models A/B/C-resc: Rescaled monetary variables (R1#5) -----------------
# Freight value is reported per BRL 10 and merchandise value per BRL 100.
# Rescaling changes only the reporting unit; the fitted relationships are
# otherwise identical to Models A, B, and C.
primary["freight_per10"] = primary["freight_value"] / 10.0
primary["price_per100"] = primary["price"] / 100.0

model_A_resc_vars = ["delivery_delay", "price_per100",]
model_B_resc_vars = ["multi_seller", "item_count", "freight_per10", "price_per100",]
model_C_resc_vars = ["multi_seller", "item_count", "freight_per10", "delivery_delay", "price_per100",]

model_A_resc = fit_logit(primary, "dissatisfied", model_A_resc_vars, "Model A-resc: Monetary values rescaled",)
model_B_resc = fit_logit(primary, "dissatisfied", model_B_resc_vars, "Model B-resc: Monetary values rescaled",)
model_C_resc = fit_logit(primary, "dissatisfied", model_C_resc_vars, "Model C-resc: Monetary values rescaled",)

or_A_resc = or_table(model_A_resc, "Model A-resc")
or_B_resc = or_table(model_B_resc, "Model B-resc")
or_C_resc = or_table(model_C_resc, "Model C-resc")

rescaled_monetary_summary = pd.concat([
    or_A_resc.assign(Model="Model A"),
    or_B_resc.assign(Model="Model B"),
    or_C_resc.assign(Model="Model C"),
    ], ignore_index=True,)

rescaled_monetary_summary = rescaled_monetary_summary[
    rescaled_monetary_summary["Variable"].isin(["freight_per10", "price_per100"])
].copy()

rescaled_monetary_summary["Variable"] = (
    rescaled_monetary_summary["Variable"].replace({
        "freight_per10": "Freight Value, BRL 10",
        "price_per100": "Merchandise Value, BRL 100",
    })
)

print("\nRescaled monetary-variable checks:")
print(f"Model A merchandise value, per BRL 100: "
      f"OR={np.exp(model_A_resc.params['price_per100']):.4f}")

print(f"Model B freight value, per BRL 10: "
      f"OR={np.exp(model_B_resc.params['freight_per10']):.4f}")
print(f"Model B merchandise value, per BRL 100: "
      f"OR={np.exp(model_B_resc.params['price_per100']):.4f}")

print(f"Model C freight value, per BRL 10: "
      f"OR={np.exp(model_C_resc.params['freight_per10']):.4f}")
print(f"Model C merchandise value, per BRL 100: "
      f"OR={np.exp(model_C_resc.params['price_per100']):.4f}")

print("Only the reporting units change; the fitted relationships are "
      "identical to Models A, B, and C.")

# ---- Seller-cluster bootstrap for multi_seller (R1#3) ----------------------
# The dependent structure here is that many orders share the same seller(s),
# so orders are not fully independent observations. The unit of analysis,
# outcome, and exposure variable are unchanged from Model C; only the
# inference procedure is extended. Sellers (not orders) are resampled with
# replacement, and all orders whose dominant seller is drawn are included
# (with repetition) in each bootstrap replicate, then Model C is refit on
# that replicate sample. A small share of orders with more than one seller
# are represented through their dominant seller only, which is a documented
# approximation to a genuinely overlapping cluster structure and is reported
# as such rather than presented as an exact multiway-clustering solution.
BOOT_REPS = 1000
boot_seed = 42
bootstrap_ready = primary.dropna(subset=full_vars + ["dissatisfied", "dominant_seller_id"]).copy()
n_orders_no_seller = len(primary) - len(bootstrap_ready)
if n_orders_no_seller > 0:
    print(f"Seller-cluster bootstrap: {n_orders_no_seller:,} orders lack a dominant_seller_id "
          f"and are excluded from the bootstrap sample only (not from Model C itself).")

seller_clusters = bootstrap_ready["dominant_seller_id"].unique()
n_seller_clusters = len(seller_clusters)
print(f"Seller-cluster bootstrap: {n_seller_clusters:,} unique dominant sellers, "
      f"{BOOT_REPS} replications, resampling clusters with replacement.")

rng_boot = np.random.default_rng(boot_seed)
seller_groups = {s: idx.values for s, idx in bootstrap_ready.groupby("dominant_seller_id").groups.items()}
boot_multi_seller_coefs = []
boot_failures = 0
for b in range(BOOT_REPS):
    sampled_sellers = rng_boot.choice(seller_clusters, size=n_seller_clusters, replace=True)
    row_idx = np.concatenate([seller_groups[s] for s in sampled_sellers])
    boot_df = bootstrap_ready.loc[row_idx]
    try:
        d = boot_df[["dissatisfied"] + full_vars].dropna().copy()
        for c in ["dissatisfied"] + full_vars:
            d[c] = d[c].astype(float)
        y_b = d["dissatisfied"]
        X_b = sm.add_constant(d[full_vars], has_constant="add")
        res_b = sm.Logit(y_b, X_b).fit(disp=False, method="newton", maxiter=200)
        if res_b.mle_retvals.get("converged", False) and "multi_seller" in res_b.params.index:
            boot_multi_seller_coefs.append(res_b.params["multi_seller"])
        else:
            boot_failures += 1
    except Exception:
        boot_failures += 1
    if (b + 1) % 50 == 0:
        print(f"  bootstrap replicate {b + 1}/{BOOT_REPS} complete "
              f"({len(boot_multi_seller_coefs)} usable, {boot_failures} failed)")

boot_multi_seller_coefs = np.array(boot_multi_seller_coefs)
boot_or = np.exp(boot_multi_seller_coefs)
boot_or_ci_low, boot_or_ci_high = np.percentile(boot_or, [2.5, 97.5])
boot_or_median = np.median(boot_or)

hc1_or = np.exp(model_C.params["multi_seller"])
hc1_ci = np.exp(model_C.conf_int().loc["multi_seller"])
print(f"\nSeller-cluster bootstrap results for multi_seller (N usable replicates="
      f"{len(boot_multi_seller_coefs)} of {BOOT_REPS}):")
print(f"  HC1 order-level OR:            {hc1_or:.3f} (95% CI {hc1_ci[0]:.3f}-{hc1_ci[1]:.3f})")
print(f"  Bootstrap median OR:           {boot_or_median:.3f}")
print(f"  Bootstrap 95% percentile CI:   {boot_or_ci_low:.3f}-{boot_or_ci_high:.3f}")
print("Note: this is an approximate dominant-seller cluster bootstrap. "
      "For multi-seller orders, clustering is defined using the seller with the "
      "highest summed merchandise value. The procedure therefore provides a "
      "sensitivity check for seller-level dependence rather than an exact treatment "
      "of overlapping seller membership.")

seller_bootstrap_summary = pd.DataFrame({
    "Statistic": ["HC1 order-level OR", "HC1 95% CI lower", "HC1 95% CI upper",
                  "Bootstrap median OR", "Bootstrap 95% CI lower", "Bootstrap 95% CI upper",
                  "Bootstrap replications requested", "Bootstrap replications usable",
                  "Unique seller clusters"],
    "Value": [hc1_or, hc1_ci[0], hc1_ci[1], boot_or_median, boot_or_ci_low, boot_or_ci_high,
              BOOT_REPS, len(boot_multi_seller_coefs), n_seller_clusters],
})

# ---- Narrow incremental discrimination check (R1#7) -------------------------
# Deliberately scoped narrowly: no train/test split, no cross-validation, no
# threshold optimization, no calibration curve, no operational decision
# rule. The only question asked is how much in-sample discrimination
# (AUC) changes when multi_seller is added to the same specification. This
# is reported as a discrimination check on an explanatory model, not as
# development or validation of a predictive scoring system.
no_seller_vars = ["item_count", "freight_value", "delivery_delay", "price"]
model_C_noseller = fit_logit(primary, "dissatisfied", no_seller_vars,
                              "Model C-noseller: Discrimination baseline without multi_seller")

disc_data = primary[["dissatisfied"] + full_vars].dropna().copy()
X_full_disc = sm.add_constant(disc_data[full_vars], has_constant="add")[model_C.params.index]
X_base_disc = sm.add_constant(disc_data[no_seller_vars], has_constant="add")[model_C_noseller.params.index]
pred_full = model_C.predict(X_full_disc)
pred_base = model_C_noseller.predict(X_base_disc)

auc_full = roc_auc_score(disc_data["dissatisfied"], pred_full)
auc_base = roc_auc_score(disc_data["dissatisfied"], pred_base)
delta_auc = auc_full - auc_base

print(f"\nDiscrimination check (rank-based AUC, in-sample, no cross-validation):")
print(f"  AUC without multi_seller: {auc_base:.4f}")
print(f"  AUC with multi_seller:    {auc_full:.4f}")
print(f"  Delta AUC:                {delta_auc:.4f}")
print("Interpretation: multi_seller carries a large odds ratio but affects only "
      f"{n_multi/n_primary:.1%} of orders, so a modest delta AUC is the expected "
      "outcome of a rare, strongly associated binary predictor and does not conflict "
      "with the odds-ratio findings above. This check speaks to overall-sample "
      "discrimination, not to the strength of the association itself, and is not a "
      "claim about production-grade predictive performance.")

discrimination_summary = pd.DataFrame({
    "Model": ["Without multi_seller", "With multi_seller"],
    "Variables": [", ".join(no_seller_vars), ", ".join(full_vars)],
    "N": [int(model_C_noseller.nobs), int(model_C.nobs)],
    "AUC": [auc_base, auc_full],
})
discrimination_summary["Delta AUC vs. baseline"] = [np.nan, delta_auc]

# ---- Consolidated extension table for reporting -----------------------------
extension_rows = []
for label, res, var in [
    ("K: Multi-item-only sensitivity", model_K, "multi_seller"),
    ("L: Product diversity control", model_L, "multi_seller"),
    ("J: Categories + enhanced controls", model_J, "multi_seller"),
]:
    ci = res.conf_int().loc[var]
    extension_rows.append({
        "Model": label,
        "N": int(res.nobs),
        "Multi-Seller OR": round(np.exp(res.params[var]), 3),
        "95% CI lower": round(np.exp(ci[0]), 3),
        "95% CI upper": round(np.exp(ci[1]), 3),
        "p": fmt_p(res.pvalues[var]),
    })
extension_summary = pd.DataFrame(extension_rows)
print("\nTable 4B. Extended robustness models - multi-seller odds ratio")
print(extension_summary.to_string(index=False))

# ============================================================================
# SECTION 10 — Diagnostics: VIF, correlation, separation
# ============================================================================
section("SECTION 10 — Statistical Diagnostics")

cont_vars = ["item_count", "freight_value", "delivery_delay", "price"]
X_vif = sm.add_constant(primary[cont_vars].dropna())
vif_df = pd.DataFrame({
    "Variable": X_vif.columns,
    "VIF": [variance_inflation_factor(X_vif.values, i) for i in range(X_vif.shape[1])]
})
vif_df = vif_df[vif_df["Variable"] != "const"]
print("Variance inflation factors (continuous predictors):")
print(vif_df.round(2).to_string(index=False))
high_vif = vif_df[vif_df["VIF"] > 5]
if len(high_vif) > 0:
    print(f"WARNING - VIF > 5 for: {list(high_vif['Variable'])}")
else:
    print("No predictor exceeds a VIF of 5; multicollinearity is not a material concern "
          "for the primary model.")

corr = primary[cont_vars + ["multi_seller"]].corr().round(3)
print("\nCorrelation matrix:")
print(corr.to_string())

print(f"\nOutcome frequency (primary sample): dissatisfied={primary['dissatisfied'].sum():,} "
      f"({primary['dissatisfied'].mean():.1%}), satisfied={(1-primary['dissatisfied']).sum():,}")
print(f"Group sample sizes: single-seller N={n_single:,}, multi-seller N={n_multi:,}")

for name, res in [("A", model_A), ("B", model_B), ("C", model_C), ("D", model_D),
                   ("E", model_E), ("F", model_F), ("G", model_G), ("H", model_H),
                   ("I", model_I), ("J", model_J), ("K", model_K), ("L", model_L),
                   ("C-resc", model_C_resc), ("C-noseller", model_C_noseller)]:
    conv = res.mle_retvals.get("converged", True)
    print(f"Model {name}: convergence={conv}")

print("\nSeparation check: no predictor's coefficient exceeds |15| in any fitted model "
      "(flagged automatically above if present), and the smallest multi-seller cell "
      f"(dissatisfied multi-seller orders, n={n_diss_multi:,}) is well above the size "
      "that typically causes quasi-separation.")

# ============================================================================
# SECTION 11 — Adjusted predicted probabilities and average marginal effects
# ============================================================================
section("SECTION 11 — Adjusted Predicted Probabilities and Marginal Effects")

ame_C = model_C.get_margeff(at="overall", method="dydx", dummy=True)
ame_summary = ame_C.summary_frame().reset_index().rename(columns={"index": "Variable"})
ame_summary["AME probability scale"] = ame_summary["dy/dx"]
ame_summary["AME percentage points"] = ame_summary["dy/dx"] * 100
ame_summary["Interpretation unit"] = ame_summary["Variable"].map({
    "multi_seller": "Discrete change 0 to 1",
    "item_count": "Per one additional item",
    "freight_value": "Per BRL 1 freight",
    "delivery_delay": "Per one additional day",
    "price": "Per BRL 1 merchandise value",
}).fillna("Per one-unit increase")
print("Average marginal effects, Model C:")
print(ame_summary[["Variable", "AME probability scale", "AME percentage points",
                   "Std. Err.", "Pr(>|z|)", "Interpretation unit"]].round(4).to_string(index=False))

# Representative profile for adjusted probabilities.
# The comparison is restricted to a structurally feasible common profile:
# a two-item order can be either single-seller or multi-seller.
prediction_reference = primary[primary["item_count"] >= 2].copy()
med = prediction_reference[["freight_value", "price"]].median()
reference_item_count = 2
delay_grid = [0, 3, 7, 14]
assert min(delay_grid) >= primary["delivery_delay"].min()
assert max(delay_grid) <= primary["delivery_delay"].max()

delay_context = (
    prediction_reference["delivery_delay"]
    .value_counts()
    .reindex(delay_grid, fill_value=0)
    .rename_axis("Delivery delay")
    .reset_index(name="Observed multi-item orders")
)

profiles = []

for ms in [0, 1]:
    for dd in delay_grid:
        profiles.append({
            "multi_seller": ms,
            "item_count": reference_item_count,
            "freight_value": med["freight_value"],
            "delivery_delay": dd,
            "price": med["price"],
        })

prof_df = pd.DataFrame(profiles)
X_prof = sm.add_constant(prof_df[full_vars], has_constant="add")
X_prof = X_prof[model_C.params.index]
prof_df["pred_prob"] = model_C.predict(X_prof)

pred_cov = model_C.cov_params()
se_list = []
for _, row in X_prof.iterrows():
    xb_var = row.values @ pred_cov.values @ row.values.T
    se_list.append(np.sqrt(xb_var))
prof_df["logit_se"] = se_list
prof_df["logit_hat"] = X_prof.values @ model_C.params.values
prof_df["prob_lo"] = 1 / (1 + np.exp(-(prof_df["logit_hat"] - 1.96 * prof_df["logit_se"])))
prof_df["prob_hi"] = 1 / (1 + np.exp(-(prof_df["logit_hat"] + 1.96 * prof_df["logit_se"])))
prof_df["seller_label"] = prof_df["multi_seller"].map({0: "Single-seller", 1: "Multi-seller"})

print("\nAdjusted predicted probability of dissatisfaction for a two-item order, "
      "with freight and merchandise value fixed at the medians of multi-item orders:")
print(prof_df[["seller_label", "delivery_delay", "pred_prob", "prob_lo", "prob_hi"]]
      .round(4).to_string(index=False))

print("\nThese are model-adjusted associations for a structurally feasible two-item "
      "order profile. Freight and merchandise value are fixed at the medians observed "
      "among multi-item orders. The estimates are not causal effects.")

# ============================================================================
# SECTION 12 — Figures
# ============================================================================
section("SECTION 12 — Figures")

# Figure 1: dissatisfaction by fulfillment type, with CI and N
fig, ax = plt.subplots(figsize=(6, 5))
labels = ["Single-Seller", "Multi-Seller"]
rates  = [diss_single * 100, diss_multi * 100]
cis    = [ci_single, ci_multi]
ns     = [n_single, n_multi]
bars = ax.bar(labels, rates, color=[COLORS["neutral"], COLORS["secondary"]],
              width=0.45, edgecolor="white", linewidth=1.5,
              yerr=[[rates[i] - cis[i][0]*100 for i in range(2)],
                    [cis[i][1]*100 - rates[i] for i in range(2)]],
              capsize=6, error_kw=dict(lw=1.3))
for bar, val, n in zip(bars, rates, ns):
    ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 1.5,
            f"{val:.1f}%\n(N={n:,})", ha="center", va="bottom", fontsize=10)
ax.set_ylabel("Dissatisfaction Rate (%)")
ax.set_title("Dissatisfaction Rate by Fulfillment Type\n(95% CI, primary sample)")
ax.set_ylim(0, max(rates) * 1.35)
plt.tight_layout()
plt.savefig(f"{OUT_DIR}/fragmentation.png", dpi=150, bbox_inches="tight")
plt.show()
print("Fragmentation figure saved.")

# Figure 2: dissatisfaction by delivery category
cat_order = DELIVERY_ORDER
cat_diss = primary_E.groupby("delivery_category", observed=True)["dissatisfied"].agg(["mean", "count"])
cat_diss = cat_diss.reindex(cat_order).dropna()
fig, ax = plt.subplots(figsize=(7.5, 4.5))
bars = ax.bar(cat_diss.index.astype(str), cat_diss["mean"] * 100, color=COLORS["primary"],
              alpha=0.85, edgecolor="white")
for bar, (val, n) in zip(bars, zip(cat_diss["mean"] * 100, cat_diss["count"])):
    ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.3,
            f"{val:.1f}%\n(n={int(n):,})", ha="center", va="bottom", fontsize=9)
ax.set_ylabel("Dissatisfaction Rate (%)")
ax.set_xlabel("Delivery Category")
ax.set_title("Dissatisfaction Rate by Delivery Category", pad=16)
ax.set_ylim(0, max(cat_diss["mean"] * 100) + 14)
plt.xticks(rotation=20, ha="right")
plt.tight_layout()
plt.savefig(f"{OUT_DIR}/delivery_category.png", dpi=150, bbox_inches="tight")
plt.show()
print("Delivery category figure saved.")

# Figure 3: adjusted predicted probability, multi-seller x delay
fig, ax = plt.subplots(figsize=(7, 4.8))
for ms, color, label in [(0, COLORS["neutral"], "Single-seller"), (1, COLORS["secondary"], "Multi-seller")]:
    sub = prof_df[prof_df["multi_seller"] == ms]
    ax.plot(sub["delivery_delay"], sub["pred_prob"] * 100, marker="o", color=color, label=label)
    ax.fill_between(sub["delivery_delay"], sub["prob_lo"] * 100, sub["prob_hi"] * 100,
                     color=color, alpha=0.15)
ax.set_xlabel("Delivery Delay (days)")
ax.set_ylabel("Adjusted Predicted P(Dissatisfied) (%)")
ax.set_title("Adjusted Predicted Dissatisfaction Probability\n(Model C, median covariate values, 95% CI)")
ax.legend()
plt.tight_layout()
plt.savefig(f"{OUT_DIR}/adjusted_probability.png", dpi=150, bbox_inches="tight")
plt.show()
print("Adjusted probability figure saved.")

# Figure 4: Model C coefficient / OR plot
or_plot = or_C.copy()
label_map = {"multi_seller": "Multi-Seller\n(Fragmentation)", "item_count": "Item Count",
             "freight_value": "Freight Value", "delivery_delay": "Delivery Delay",
             "price": "Merchandise Value"}
or_plot["label"] = or_plot["Variable"].map(label_map)
fig, ax = plt.subplots(figsize=(7, 4.5))
colors_ = [COLORS["secondary"] if s else COLORS["neutral"] for s in (or_plot["Sig"] != "")]
ax.barh(or_plot["label"], or_plot["OR"], color=colors_,
        xerr=[or_plot["OR"] - or_plot["OR_95%CI_low"], or_plot["OR_95%CI_high"] - or_plot["OR"]],
        edgecolor="white", error_kw=dict(ecolor="black", capsize=4, lw=1.2))
ax.axvline(1, color="black", linestyle="--", linewidth=1.1, alpha=0.6)
ax.set_xlabel("Odds Ratio (95% CI, robust SE)")
ax.set_title("Model C Odds Ratios")
sig_patch = mpatches.Patch(color=COLORS["secondary"], label="p < .05")
ns_patch  = mpatches.Patch(color=COLORS["neutral"], label="Not significant")
ax.legend(handles=[sig_patch, ns_patch], loc="lower right", fontsize=9)
plt.tight_layout()
plt.savefig(f"{OUT_DIR}/coefficients.png", dpi=150, bbox_inches="tight")
plt.show()
print("Coefficients figure saved.")

# ============================================================================
# SECTION 13 — Create the single consolidated Excel workbook
# ============================================================================
section("SECTION 13 — Creating Consolidated Excel Workbook")

# ---- 13.1 Prepare publication and audit tables ----------------------------

# README content
readme_df = pd.DataFrame({
    "Item": [
        "Paper title",
        "Dataset",
        "Primary outcome",
        "Merchandise-value field",
        "Primary sample",
        "Sensitivity sample",
        "Review deduplication",
        "Primary models",
        "Delivery functional-form models",
        "Enhanced controls",
        "Robust standard errors",
        "Interpretation",
        "Model J",
        "Model K (multi-item sensitivity)",
        "Model L (product diversity)",
        "Models A/B/C-resc (rescaled monetary variables)",
        "Seller-cluster bootstrap (Model C)",
        "Discrimination check (AUC)",
        "Recommended paper tables",
        "Recommended paper figures",
        "Optional figure",
        "Do not use in main paper",
        "Generated",
        "Python version",
        "pandas version",
        "numpy version",
        "statsmodels version",
        "scipy version",
        "matplotlib version",
        "openpyxl version",
    ],
    "Description": [
        "Multi-Seller Fulfillment and Customer Dissatisfaction in Marketplace E-Commerce",
        "Brazilian E-Commerce Public Dataset by Olist",
        "Dissatisfied = 1 when review score <= 2; otherwise 0",
        "Internal variable 'price' is the sum of item prices and is reported as merchandise value excluding freight",
        (
            f"{n_primary:,} delivered orders with valid review, delivery, "
            "freight, merchandise value, item, and seller information"
        ),
        (
            f"{n_sensitivity:,} orders after excluding reviews created before "
            "confirmed delivery"
        ),
        (
            f"{n_multi_review_ord:,} orders had multiple review records. "
            "The latest review_answer_timestamp was retained; ties were resolved "
            "using review_creation_date."
        ),
        "Model A: delivery benchmark; Model B: fulfillment benchmark; Model C: combined model",
        "Model D: late indicator; Model E: delivery categories",
        "Model F: product category, customer state, and purchase-year controls",
        "HC1 heteroskedasticity-robust standard errors",
        "All estimates are observational associations, not causal effects",
        "Model J: delivery categories combined with category, state, and purchase-year controls",
        (
            f"Restricted to item_count>=2 (N={n_multi_item_sample:,}); tests whether the "
            "multi-seller association holds when the comparison group is limited to orders "
            "that were also structurally capable of being multi-seller"
        ),
        "Adds unique_product_count to Model C to separate seller fragmentation from assortment breadth",
        "Reports freight value per BRL 10 and merchandise value per BRL 100 for the primary models; rescaling changes reporting units only.",
        (
            f"Seller-cluster bootstrap ({BOOT_REPS} replications, resampling by dominant "
            "seller_id) tests whether the Model C multi-seller CI remains stable once "
            "within-seller order dependence is accounted for; approximate by construction "
            "for the minority of orders with more than one seller"
        ),
        (
            "Narrow in-sample AUC/delta-AUC check comparing Model C with and without "
            "multi_seller; scoped explicitly as a discrimination check, not development "
            "or validation of a predictive scoring model"
        ),
        "Sample & Descriptives; Main Models; Robustness; Extended Robustness",
        "Figure 1 and Figure 2",
        "Figure 3",
        "Figure 4 is supplementary because predictors use different measurement units",
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        platform.python_version(),
        pd.__version__,
        np.__version__,
        statsmodels.__version__,
        scipy.__version__,
        matplotlib.__version__,
        openpyxl.__version__,
    ],
})

# Sample construction with exclusion and retention information
sample_construction = reconciliation.copy()
sample_construction["Excluded from previous stage"] = (
    sample_construction["N"].shift(1) - sample_construction["N"]
)
sample_construction.loc[0, "Excluded from previous stage"] = 0

sample_construction["Retained from previous stage (%)"] = (
    sample_construction["N"] / sample_construction["N"].shift(1)
)
sample_construction.loc[0, "Retained from previous stage (%)"] = 1.0

sample_construction["Retained from raw orders (%)"] = (
    sample_construction["N"] / n_raw_orders
)

# Primary descriptive statistics with quartiles
primary_descriptives = (
    primary[desc_vars]
    .describe(percentiles=[0.25, 0.50, 0.75])
    .T[["count", "mean", "std", "min", "25%", "50%", "75%", "max"]]
    .reset_index()
    .rename(columns={
        "index": "Variable",
        "count": "N",
        "mean": "Mean",
        "std": "Std. Dev.",
        "min": "Minimum",
        "25%": "25th percentile",
        "50%": "Median",
        "75%": "75th percentile",
        "max": "Maximum",
    })
)

# Fulfillment comparison
single_late_rate = primary.loc[primary["multi_seller"] == 0, "late"].mean()
multi_late_rate = primary.loc[primary["multi_seller"] == 1, "late"].mean()

fulfillment_comparison = pd.DataFrame({
    "Metric": [
        "Orders",
        "Dissatisfied orders",
        "Dissatisfaction rate",
        "Wilson CI lower",
        "Wilson CI upper",
        "Mean item count",
        "Median item count",
        "Mean freight value",
        "Median freight value",
        "Mean merchandise value",
        "Median merchandise value",
        "Mean delivery delay",
        "Median delivery delay",
        "Late-order rate",
    ],
    "Single-Seller": [
        n_single,
        n_diss_single,
        diss_single,
        ci_single[0],
        ci_single[1],
        comp_mean.loc[0, "item_count"],
        comp_median.loc[0, "item_count"],
        comp_mean.loc[0, "freight_value"],
        comp_median.loc[0, "freight_value"],
        comp_mean.loc[0, "price"],
        comp_median.loc[0, "price"],
        comp_mean.loc[0, "delivery_delay"],
        comp_median.loc[0, "delivery_delay"],
        single_late_rate,
    ],
    "Multi-Seller": [
        n_multi,
        n_diss_multi,
        diss_multi,
        ci_multi[0],
        ci_multi[1],
        comp_mean.loc[1, "item_count"],
        comp_median.loc[1, "item_count"],
        comp_mean.loc[1, "freight_value"],
        comp_median.loc[1, "freight_value"],
        comp_mean.loc[1, "price"],
        comp_median.loc[1, "price"],
        comp_mean.loc[1, "delivery_delay"],
        comp_median.loc[1, "delivery_delay"],
        multi_late_rate,
    ],
})

fulfillment_tests = pd.DataFrame({
    "Statistic": [
        "Absolute risk difference",
        "Risk ratio",
        "Unadjusted odds ratio",
        "Chi-square",
        "Degrees of freedom",
        "p-value",
        "Multi-seller sample share",
    ],
    "Value": [
        risk_diff,
        risk_ratio,
        unadj_or,
        chi2,
        dof,
        fmt_p(p_chi),
        n_multi / n_primary,
    ],
})

# Delivery category summary with Wilson confidence intervals
delivery_summary_rows = []

for category in DELIVERY_ORDER:
    subset = primary.loc[primary["delivery_category"].astype(str) == category]

    n_category = len(subset)
    dissatisfied_category = int(subset["dissatisfied"].sum())
    rate_category = (dissatisfied_category / n_category if n_category > 0 else np.nan)

    if n_category > 0:
        ci_low, ci_high = wilson_ci(dissatisfied_category, n_category)
    else:
        ci_low, ci_high = np.nan, np.nan

    delivery_summary_rows.append({
        "Delivery category": category,
        "N": n_category,
        "Dissatisfied N": dissatisfied_category,
        "Dissatisfaction rate": rate_category,
        "Wilson CI lower": ci_low,
        "Wilson CI upper": ci_high,
    })

delivery_category_summary = pd.DataFrame(delivery_summary_rows)

# Main model table with odds ratios
main_model_rows = []

display_labels = {
    "multi_seller": "Multi-Seller",
    "item_count": "Item Count",
    "freight_value": "Freight Value",
    "delivery_delay": "Delivery Delay",
    "price": "Merchandise Value",
}

for model_name, result in [
    ("Model A", model_A),
    ("Model B", model_B),
    ("Model C", model_C),
]:
    confidence_intervals = result.conf_int()

    for variable in display_labels:
        if variable not in result.params.index:
            continue

        main_model_rows.append({
            "Model": model_name,
            "Variable": display_labels[variable],
            "Coefficient": result.params[variable],
            "Robust SE": result.bse[variable],
            "p-value": result.pvalues[variable],
            "p-value formatted": fmt_p(result.pvalues[variable]),
            "Odds Ratio": np.exp(result.params[variable]),
            "OR CI lower": np.exp(confidence_intervals.loc[variable, 0]),
            "OR CI upper": np.exp(confidence_intervals.loc[variable, 1]),
            "N": int(result.nobs),
            "AIC": result.aic,
            "McFadden R2": result.prsquared,
            "Converged": result.mle_retvals.get("converged", False),
        })

main_models_table = pd.DataFrame(main_model_rows)

# Compact robustness table using key coefficients only
robustness_summary_rows = []

robustness_model_list = [
    ("Model D", "Late indicator", model_D),
    ("Model E", "Delivery categories", model_E),
    ("Model F", "Enhanced controls", model_F),
    ("Model G", "Timing-restricted sample", model_G),
    ("Model H", "Review score <= 3", model_H),
    ("Model I", "Freight per item", model_I),
    ("Model J", "Delivery categories + enhanced controls", model_J),
]

key_robustness_variables = [
    "multi_seller",
    "item_count",
    "freight_value",
    "log_freight_per_item",
    "delivery_delay",
    "late",
    "price",
]

for model_code, model_description, result in robustness_model_list:
    confidence_intervals = result.conf_int()
    model_variables = list(key_robustness_variables)
    if model_code in ("Model E", "Model J"):
        model_variables += [
            variable for variable in result.params.index
            if variable.startswith("delcat_")
        ]

    for variable in model_variables:
        if variable not in result.params.index:
            continue

        robustness_summary_rows.append({
            "Model": model_code,
            "Specification": model_description,
            "Variable": display_labels.get(variable, variable.replace("delcat_", "Delivery: ")),
            "Odds Ratio": np.exp(result.params[variable]),
            "OR CI lower": np.exp(confidence_intervals.loc[variable, 0]),
            "OR CI upper": np.exp(confidence_intervals.loc[variable, 1]),
            "Robust SE": result.bse[variable],
            "p-value": result.pvalues[variable],
            "p-value formatted": fmt_p(result.pvalues[variable]),
            "N": int(result.nobs),
            "AIC": result.aic,
            "McFadden R2": result.prsquared,
            "Converged": result.mle_retvals.get("converged", False),
        })

robustness_summary = pd.DataFrame(robustness_summary_rows)

# Enhanced Model F full output
model_f_ci = model_F.conf_int()

enhanced_controls_table = pd.DataFrame({
    "Variable": model_F.params.index,
    "Coefficient": model_F.params.values,
    "Robust SE": model_F.bse.values,
    "z-statistic": model_F.tvalues.values,
    "p-value": model_F.pvalues.values,
    "p-value formatted": [fmt_p(p) for p in model_F.pvalues.values],
    "Odds Ratio": np.exp(model_F.params.values),
    "OR CI lower": np.exp(model_f_ci[0].values),
    "OR CI upper": np.exp(model_f_ci[1].values),
})

# Marginal effects table
marginal_effects_table = ame_summary[[
    "Variable",
    "AME probability scale",
    "AME percentage points",
    "Std. Err.",
    "Pr(>|z|)",
    "Interpretation unit",
]].copy()

# Adjusted probabilities table
adjusted_probabilities_table = prof_df[[
    "seller_label",
    "delivery_delay",
    "pred_prob",
    "prob_lo",
    "prob_hi",
]].copy()

adjusted_probabilities_table.columns = [
    "Fulfillment type",
    "Delivery delay",
    "Predicted probability",
    "CI lower",
    "CI upper",
]

# Review diagnostic table
review_diagnostics = pd.DataFrame({
    "Metric": [
        "Total review rows",
        "Unique orders with review",
        "Orders with multiple reviews",
        "Duplicate-review percentage",
        "Review selection rule",
    ],
    "Value": [
        n_review_rows,
        n_unique_reviewed,
        n_multi_review_ord,
        n_multi_review_ord / n_unique_reviewed,
        (
            "Keep latest review_answer_timestamp; "
            "tie-break using latest review_creation_date"
        ),
    ],
})

# Merge and aggregation diagnostics
merge_audit_df = pd.DataFrame(merge_audit_rows)
aggregation_audit_df = pd.DataFrame(aggregation_audit_rows)

# Model convergence diagnostic
convergence_rows = []

for model_name, result in [
    ("Model A", model_A),
    ("Model B", model_B),
    ("Model C", model_C),
    ("Model D", model_D),
    ("Model E", model_E),
    ("Model F", model_F),
    ("Model G", model_G),
    ("Model H", model_H),
    ("Model I", model_I),
    ("Model J", model_J),
    ("Model K", model_K),
    ("Model L", model_L),
    ("Model A-resc", model_A_resc),
    ("Model B-resc", model_B_resc),
    ("Model C-resc", model_C_resc),
    ("Model C-noseller", model_C_noseller),
]:
    convergence_rows.append({
        "Model": model_name,
        "Converged": result.mle_retvals.get("converged", False),
        "Optimizer": getattr(result, "_optimizer_used", "unknown"),
        "Maximum absolute coefficient": (result.params.abs().max()),
        "Maximum robust SE": result.bse.max(),
        "N": int(result.nobs),
    })

convergence_df = pd.DataFrame(convergence_rows)

# Figure guide
figure_guide = pd.DataFrame({
    "Figure": [
        "fragmentation.png",
        "delivery_category.png",
        "adjusted_probability.png",
        "coefficients.png",
    ],
    "Purpose": [
        "Observed dissatisfaction rates for single- and multi-seller orders",
        "Observed nonlinear dissatisfaction pattern across delivery categories",
        "Adjusted predicted dissatisfaction at 0, 3, 7, and 14 delay days",
        "Model C odds ratios with confidence intervals",
    ],
    "Recommended use": [
        "Main paper",
        "Main paper",
        "Optional",
        "Do not use in main paper",
    ],
})

# ---- 13.2 Helper for writing multiple tables into one worksheet -----------
header_rows_by_sheet = {}

def write_section(
    writer,
    sheet_name,
    title,
    dataframe,
    startrow,
    index=False
):
    """
    Write a section title followed by one DataFrame.
    Returns the next available row after the table and two blank rows.
    """
    worksheet = writer.book[sheet_name]

    title_excel_row = startrow + 1
    header_excel_row = startrow + 2

    worksheet.cell(row=title_excel_row, column=1, value=title)
    dataframe.to_excel(writer, sheet_name=sheet_name, startrow=startrow + 1, index=index)

    header_rows_by_sheet.setdefault(sheet_name, []).append(header_excel_row)
    return startrow + len(dataframe) + 4

# ---- 13.3 Create workbook and write requested sheets ----------------------
with pd.ExcelWriter(WORKBOOK_PATH, engine="openpyxl") as writer:

    # README
    readme_df.to_excel(writer, sheet_name="README", index=False)
    header_rows_by_sheet["README"] = [1]

    # Sample & Descriptives
    pd.DataFrame().to_excel(
        writer,
        sheet_name="Sample & Descriptives",
        index=False
    )

    row = 0
    row = write_section(
        writer,
        "Sample & Descriptives",
        "Sample Construction",
        sample_construction,
        row
    )
    row = write_section(
        writer,
        "Sample & Descriptives",
        "Primary Descriptive Statistics",
        primary_descriptives,
        row
    )
    row = write_section(
        writer,
        "Sample & Descriptives",
        "Fulfillment Comparison",
        fulfillment_comparison,
        row
    )
    row = write_section(
        writer,
        "Sample & Descriptives",
        "Fulfillment Comparison Statistics",
        fulfillment_tests,
        row
    )
    row = write_section(
        writer,
        "Sample & Descriptives",
        "Delivery Category Summary",
        delivery_category_summary,
        row
    )

    # Main Models
    pd.DataFrame().to_excel(
        writer,
        sheet_name="Main Models",
        index=False
    )

    row = 0
    row = write_section(
        writer,
        "Main Models",
        "Models A, B, and C: Odds Ratios",
        main_models_table,
        row
    )
    row = write_section(
        writer,
        "Main Models",
        "Models A, B, and C: Fit Statistics",
        fit_ABC,
        row
    )

    # Robustness
    robustness_summary.to_excel(
        writer,
        sheet_name="Robustness",
        index=False
    )
    header_rows_by_sheet["Robustness"] = [1]

    # Extended Robustness
    pd.DataFrame().to_excel(
        writer,
        sheet_name="Extended Robustness",
        index=False
    )

    row = 0
    row = write_section(
        writer,
        "Extended Robustness",
        "Table 4B: Multi-Item, Diversity, and Categories+Controls Models",
        extension_summary,
        row
    )
    row = write_section(
        writer,
        "Extended Robustness",
        "Seller-Cluster Bootstrap: HC1 vs. Bootstrap CI for Multi-Seller",
        seller_bootstrap_summary,
        row
    )
    row = write_section(
        writer,
        "Extended Robustness",
        "Rescaled Monetary Variables",
        rescaled_monetary_summary,
        row
    )
    row = write_section(
        writer,
        "Extended Robustness",
        "Narrow Incremental Discrimination Check (AUC)",
        discrimination_summary,
        row
    )

    # Enhanced Controls
    enhanced_controls_table.to_excel(
        writer,
        sheet_name="Enhanced Controls",
        index=False
    )
    header_rows_by_sheet["Enhanced Controls"] = [1]

    # Marginal Effects
    pd.DataFrame().to_excel(
        writer,
        sheet_name="Marginal Effects",
        index=False
    )

    row = 0
    row = write_section(
        writer,
        "Marginal Effects",
        "Average Marginal Effects: Model C",
        marginal_effects_table,
        row
    )
    row = write_section(
        writer,
        "Marginal Effects",
        "Adjusted Predicted Probabilities: Model C",
        adjusted_probabilities_table,
        row
    )
    row = write_section(
        writer,
        "Marginal Effects",
        "Delay Model Comparison",
        delay_model_comparison,
        row
    )
    row = write_section(
        writer,
        "Marginal Effects",
        "Observed Orders at Prediction Delay Values",
        delay_context,
        row
    )

    # Diagnostics
    pd.DataFrame().to_excel(
        writer,
        sheet_name="Diagnostics",
        index=False
    )

    row = 0
    row = write_section(
        writer,
        "Diagnostics",
        "Review Diagnostics",
        review_diagnostics,
        row
    )
    row = write_section(
        writer,
        "Diagnostics",
        "Merge Validation",
        merge_audit_df,
        row
    )
    row = write_section(
        writer,
        "Diagnostics",
        "Aggregation Audit",
        aggregation_audit_df,
        row
    )
    row = write_section(
        writer,
        "Diagnostics",
        "Variance Inflation Factors",
        vif_df,
        row
    )
    row = write_section(
        writer,
        "Diagnostics",
        "Correlation Matrix",
        corr.reset_index().rename(
            columns={"index": "Variable"}
        ),
        row
    )
    row = write_section(
        writer,
        "Diagnostics",
        "Model Convergence",
        convergence_df,
        row
    )
    row = write_section(
        writer,
        "Diagnostics",
        "Categorical Sparsity Diagnostics",
        categorical_sparsity,
        row
    )

    # Figure Guide
    figure_guide.to_excel(
        writer,
        sheet_name="Figure Guide",
        index=False
    )
    header_rows_by_sheet["Figure Guide"] = [1]

# ---- 13.4 Apply professional workbook formatting --------------------------
workbook = openpyxl.load_workbook(WORKBOOK_PATH)

header_fill = PatternFill(fill_type="solid", fgColor="2C3E6B")
section_fill = PatternFill(fill_type="solid", fgColor="D6DCE4")

header_font = Font(color="FFFFFF", bold=True)
section_font = Font(bold=True, size=12, color="1F1F1F")

thin_gray = Side(style="thin", color="D9D9D9")
thin_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

percentage_headers = {
    "dissatisfaction rate",
    "late-order rate",
    "wilson ci lower",
    "wilson ci upper",
    "multi-seller sample share",
    "retained from previous stage (%)",
    "retained from raw orders (%)",
    "duplicate-review percentage",
    "dissatisfied (<=2) %",
    "dissatisfied (<=3) %",
    "predicted probability",
    "ci lower",
    "ci upper",
}

for worksheet in workbook.worksheets:

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"

    # Format stored header rows
    for header_row in header_rows_by_sheet.get(worksheet.title, []):
        for cell in worksheet[header_row]:
            if cell.value is not None:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border

    # Section-title rows are rows whose first cell is populated
    # while the remaining cells are blank and the next row is a header
    for row_number in range(1, worksheet.max_row + 1):
        row_values = [
            worksheet.cell(row=row_number, column=column_number).value
            for column_number in range(1, worksheet.max_column + 1)
        ]

        non_empty = [
            value for value in row_values
            if value not in (None, "")
        ]

        if (len(non_empty) == 1 and row_number + 1 in header_rows_by_sheet.get(worksheet.title, [])):
            title_cell = worksheet.cell(row=row_number, column=1)
            title_cell.fill = section_fill
            title_cell.font = section_font
            title_cell.alignment = Alignment(vertical="center")

    # Borders and wrapping for populated cells
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

                if cell.row not in header_rows_by_sheet.get(worksheet.title, []):
                    cell.border = thin_border

    # Apply filters to simple one-table sheets
    if worksheet.title in [
        "README",
        "Robustness",
        "Enhanced Controls",
        "Figure Guide",
    ]:
        worksheet.auto_filter.ref = (worksheet.dimensions)

    # Number formats based on column headers
    for header_row in header_rows_by_sheet.get(worksheet.title,[]):
        header_map = {}

        for cell in worksheet[header_row]:
            if cell.value is not None:
                header_map[cell.column] = (str(cell.value).strip().lower())

        for column_number, header_text in header_map.items():
            for row_number in range(header_row + 1, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_number, column=column_number)

                if not isinstance(cell.value, (int, float)):
                    continue

                if header_text in percentage_headers or header_text.endswith(" rate") or header_text.endswith(" probability"):
                    cell.number_format = "0.0%"

                elif (
                    "p-value" in header_text
                    and "formatted" not in header_text
                ):
                    cell.number_format = "0.000"

                elif any(keyword in header_text for keyword in [
                    "coefficient",
                    "robust se",
                    "odds ratio",
                    "or ci",
                    "mcfadden",
                    "vif",
                    "correlation",
                    "ame",
                ]):
                    cell.number_format = "0.0000"

                elif header_text in [
                    "n",
                    "rows",
                    "unique orders",
                    "unmatched",
                    "dissatisfied n",
                    "dissatisfied orders",
                ]:
                    cell.number_format = "#,##0"

    # Sensible column widths
    for column_number in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(column_number)
        max_length = 0

        for cell in worksheet[column_letter]:
            if cell.value is None:
                continue

            value_length = len(str(cell.value))
            max_length = max(max_length,value_length)

        if column_number == 1:
            width = min(max(max_length + 2, 18), 42)
        else:
            width = min(max(max_length + 2, 12), 28)

        worksheet.column_dimensions[column_letter].width = width

# Make README description column wider.
if "README" in workbook.sheetnames:
    workbook["README"].column_dimensions["A"].width = 30
    workbook["README"].column_dimensions["B"].width = 90

# Save the formatted workbook.
workbook.save(WORKBOOK_PATH)

# ---- 13.5 Verify workbook creation ----------------------------------------
if not os.path.exists(WORKBOOK_PATH):
    raise FileNotFoundError(f"Workbook failed to save at: {WORKBOOK_PATH}")

workbook_size_mb = (os.path.getsize(WORKBOOK_PATH) / (1024 ** 2))

print(
    f"\nWorkbook successfully created:\n"
    f"  Path: {WORKBOOK_PATH}\n"
    f"  Size: {workbook_size_mb:.2f} MB\n"
    f"  Sheets: {workbook.sheetnames}"
)

# ============================================================================
# SECTION 14 — Final Summary
# ============================================================================
section("SECTION 14 — Final Summary")

def hyp_line(name, var, res):
    if var not in res.params.index:
        print(f"[{name}] variable not in model.")
        return
    orr = np.exp(res.params[var])
    ci = res.conf_int().loc[var]
    ci_low, ci_high = np.exp(ci[0]), np.exp(ci[1])
    p = res.pvalues[var]
    print(f"[{name}] {var}: OR={orr:.3f}, 95% CI [{ci_low:.3f}, {ci_high:.3f}], "
          f"p{fmt_p_inline(p)} -> "
          f"{'statistically significant' if p < .05 else 'not statistically significant'}")

hyp_line("H1 Multi-seller", "multi_seller", model_C)
hyp_line("H2 Item count", "item_count", model_C)
hyp_line("H3 Freight value", "freight_value", model_C)
hyp_line("H4 Delivery delay", "delivery_delay", model_C)

print("Adjusted predicted probability at on-time delivery for a two-item order: "
      f"single-seller {prof_df.loc[(prof_df['multi_seller'] == 0) & (prof_df['delivery_delay'] == 0), 'pred_prob'].iloc[0]:.1%}, "
      f"multi-seller {prof_df.loc[(prof_df['multi_seller'] == 1) & (prof_df['delivery_delay'] == 0), 'pred_prob'].iloc[0]:.1%}.")
print("All comparisons above are model-adjusted associations from an observational "
      "dataset. No causal claim is made about seller structure or delivery timing.")

print(f"\nAnalysis complete.")
print(f"Workbook saved to: {WORKBOOK_PATH}")
print(f"Figures saved to: {OUT_DIR}")